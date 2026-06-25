"""
Gerador do artefato: FS - Roteiro de Testes.xlsx
Execução: python gerar_roteiro_testes.py
Dependência: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ---------------------------------------------------------------------------
# Paleta de cores
# ---------------------------------------------------------------------------
COR_VERDE_ESCURO = "114D25"
COR_AZUL         = "1E63FF"
COR_CINZA_CLARO  = "F2F2F2"
COR_AMARELO      = "FFF2CC"
COR_VERDE_CLARO  = "E2EFDA"
COR_VERMELHO     = "FFE0E0"
COR_BRANCO       = "FFFFFF"

borda = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def cabecalho_fill(cor):
    return PatternFill("solid", fgColor=cor)

def fonte_branca_negrito():
    return Font(color="FFFFFF", bold=True, name="Calibri", size=10)

def fonte_normal():
    return Font(name="Calibri", size=10)

def alinhar(horizontal="left", vertical="center", wrap=True):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def aplicar_borda_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = borda

# ---------------------------------------------------------------------------
# Dados: Roteiro de Testes
# ---------------------------------------------------------------------------
COLUNAS_ROTEIRO = [
    "ID", "UC", "Nome do Caso de Teste", "Tipo",
    "Pré-condições", "Passos", "Dados de Entrada",
    "Resultado Esperado", "Resultado Obtido", "Status", "Observações"
]

CASOS_TESTE = [
    # -----------------------------------------------------------------------
    # UC-02 — Fazer Login
    # -----------------------------------------------------------------------
    ("CT-001", "UC-02", "Login com e-mail e senha válidos", "Funcional",
     "Sistema disponível; usuário cadastrado com e-mail 'teste@email.com' e senha '123'",
     "1. Acessar a tela de Login\n2. Inserir e-mail: teste@email.com\n3. Inserir senha: 123\n4. Clicar em 'Entrar'",
     "email=teste@email.com | senha=123",
     "Sistema exibe mensagem 'Login realizado com sucesso!' e redireciona para /home", "", "—", ""),

    ("CT-002", "UC-02", "Login com senha incorreta", "Funcional",
     "Sistema disponível; usuário cadastrado com e-mail 'teste@email.com'",
     "1. Acessar a tela de Login\n2. Inserir e-mail: teste@email.com\n3. Inserir senha incorreta\n4. Clicar em 'Entrar'",
     "email=teste@email.com | senha=senhaerrada",
     "Sistema exibe 'E-mail ou senha incorretos'; permanece na tela de Login", "", "—", ""),

    ("CT-003", "UC-02", "Login com e-mail não cadastrado", "Funcional",
     "Sistema disponível; e-mail 'naocadastrado@email.com' não existe na base",
     "1. Acessar a tela de Login\n2. Inserir e-mail não cadastrado\n3. Inserir qualquer senha\n4. Clicar em 'Entrar'",
     "email=naocadastrado@email.com | senha=123",
     "Sistema exibe 'E-mail ou senha incorretos'; permanece na tela de Login", "", "—", ""),

    ("CT-004", "UC-02", "Login com campos em branco", "Funcional",
     "Sistema disponível",
     "1. Acessar a tela de Login\n2. Deixar campos e-mail e senha em branco\n3. Clicar em 'Entrar'",
     "email='' | senha=''",
     "Navegador/sistema impede envio e exibe validação de campo obrigatório", "", "—", ""),

    ("CT-005", "UC-02", "Login com e-mail em formato inválido", "Funcional",
     "Sistema disponível",
     "1. Acessar a tela de Login\n2. Inserir e-mail sem '@'\n3. Inserir senha\n4. Clicar em 'Entrar'",
     "email=email-sem-arroba | senha=123",
     "Campo tipo 'email' do HTML5 exibe erro de formato antes do envio", "", "—", ""),

    ("CT-006", "UC-02", "Tempo de resposta do login < 2 segundos", "Não Funcional",
     "Sistema disponível; credenciais válidas",
     "1. Registrar timestamp inicial\n2. Enviar requisição POST /auth/login\n3. Registrar timestamp final",
     "email=teste@email.com | senha=123",
     "Resposta retornada em menos de 2000ms", "", "—", "Requisito de desempenho"),

    ("CT-007", "UC-02", "Múltiplas tentativas de login com falha não travam o sistema", "Não Funcional",
     "Sistema disponível",
     "1. Executar 10 tentativas de login com credenciais inválidas consecutivas\n2. Verificar disponibilidade do sistema",
     "email=naocadastrado@email.com | senha=errada (10x)",
     "Sistema permanece responsivo após 10 falhas; retorna 'Credenciais inválidas' em cada tentativa", "", "—", ""),

    # -----------------------------------------------------------------------
    # UC-01 — Manter Loja
    # -----------------------------------------------------------------------
    ("CT-008", "UC-01", "Cadastrar loja com dados válidos", "Funcional",
     "Usuário autenticado com perfil de lojista",
     "1. Acessar 'Cadastrar Loja'\n2. Preencher CNPJ: 12.345.678/0001-99\n3. Preencher nome: Mercado Teste\n4. Preencher descrição\n5. Salvar",
     "cnpj=12.345.678/0001-99 | nome=Mercado Teste | descricao=Loja de alimentos perto do prazo",
     "Loja cadastrada com sucesso; exibida na listagem", "", "—", ""),

    ("CT-009", "UC-01", "Cadastrar loja com CNPJ inválido", "Funcional",
     "Usuário autenticado com perfil de lojista",
     "1. Acessar 'Cadastrar Loja'\n2. Preencher CNPJ inválido\n3. Tentar salvar",
     "cnpj=00.000.000/0000-00",
     "Sistema exibe mensagem de CNPJ inválido; cadastro não é realizado", "", "—", ""),

    ("CT-010", "UC-01", "Cadastrar loja com nome duplicado", "Funcional",
     "Loja 'Mercado Teste' já cadastrada no sistema",
     "1. Acessar 'Cadastrar Loja'\n2. Inserir o mesmo nome de loja já existente\n3. Salvar",
     "nome=Mercado Teste (duplicado)",
     "Sistema exibe aviso de nome já utilizado ou permite duplicata (comportamento a definir)", "", "—", ""),

    ("CT-011", "UC-01", "Editar informações da loja", "Funcional",
     "Loja id=1 cadastrada; usuário autenticado como dono",
     "1. Acessar 'Minhas Lojas'\n2. Selecionar loja\n3. Editar descrição\n4. Salvar",
     "id=1 | nova_descricao=Descrição atualizada",
     "Alterações salvas; loja exibida com novos dados", "", "—", ""),

    ("CT-012", "UC-01", "Desativar loja ativa", "Funcional",
     "Loja id=1 ativa; usuário autenticado como dono",
     "1. Acessar 'Minhas Lojas'\n2. Selecionar loja ativa\n3. Clicar em 'Desativar'\n4. Confirmar",
     "id=1 | status=Ativo → Inativo",
     "Loja desativada; não exibida para clientes em buscas", "", "—", ""),

    ("CT-013", "UC-01", "Visualizar loja existente", "Funcional",
     "Loja id=1 cadastrada",
     "1. Navegar para página da loja id=1",
     "id=1",
     "Página da loja exibida com nome, descrição e produtos", "", "—", ""),

    ("CT-014", "UC-01", "Cadastrar loja com campos obrigatórios em branco", "Funcional",
     "Usuário autenticado com perfil de lojista",
     "1. Acessar 'Cadastrar Loja'\n2. Deixar campo 'Nome' em branco\n3. Tentar salvar",
     "nome='' | cnpj=12.345.678/0001-99",
     "Sistema exibe erro de campo obrigatório; cadastro não realizado", "", "—", ""),

    ("CT-015", "UC-01", "Tempo de carregamento da página da loja < 3 segundos", "Não Funcional",
     "Loja com 20 produtos cadastrados",
     "1. Navegar para página da loja\n2. Medir tempo de carregamento",
     "id=1 (20 produtos)",
     "Página carregada em menos de 3000ms", "", "—", "Requisito de desempenho"),

    # -----------------------------------------------------------------------
    # UC-03 — Manter Produtos
    # -----------------------------------------------------------------------
    ("CT-016", "UC-03", "Cadastrar produto com validade futura válida", "Funcional",
     "Usuário autenticado; loja id=1 ativa",
     "1. Acessar 'Meus Produtos'\n2. Clicar em 'Adicionar Produto'\n3. Preencher nome, preço, validade futura\n4. Salvar",
     "nome=Iogurte Natural | preco=1.99 | validade=2026-06-16",
     "Produto cadastrado e listado na loja", "", "—", ""),

    ("CT-017", "UC-03", "Cadastrar produto com data de validade passada", "Funcional",
     "Usuário autenticado; loja id=1 ativa",
     "1. Acessar 'Meus Produtos'\n2. Inserir validade anterior à data atual\n3. Tentar salvar",
     "nome=Queijo Fatiado | preco=0.50 | validade=2026-06-12",
     "Sistema exibe aviso de produto vencido ou bloqueia o cadastro (comportamento a definir)", "", "—", ""),

    ("CT-018", "UC-03", "Editar preço de produto existente", "Funcional",
     "Produto id=1 cadastrado; usuário autenticado como dono da loja",
     "1. Selecionar produto id=1\n2. Alterar preço\n3. Salvar",
     "id=1 | novo_preco=2.49",
     "Preço atualizado e exibido corretamente", "", "—", ""),

    ("CT-019", "UC-03", "Remover produto da loja", "Funcional",
     "Produto id=1 cadastrado; sem pedidos pendentes associados",
     "1. Selecionar produto id=1\n2. Clicar em 'Remover'\n3. Confirmar exclusão",
     "id=1",
     "Produto removido; não exibido na loja", "", "—", ""),

    ("CT-020", "UC-03", "Cadastrar produto sem foto", "Funcional",
     "Usuário autenticado; loja ativa",
     "1. Preencher dados do produto sem adicionar imagem\n2. Tentar salvar",
     "nome=Produto Sem Foto | preco=0.99 | validade=2026-07-01 | foto=null",
     "Sistema cadastra produto com imagem padrão OU exige foto (comportamento a definir)", "", "—", ""),

    ("CT-021", "UC-03", "Cadastrar produto com preço zerado", "Funcional",
     "Usuário autenticado; loja ativa",
     "1. Inserir preço = 0,00\n2. Salvar",
     "preco=0.00",
     "Sistema exibe validação de preço inválido ou permite (regra de negócio a confirmar)", "", "—", ""),

    ("CT-022", "UC-03", "Listar todos os produtos de uma loja", "Funcional",
     "Loja id=1 com 2 produtos cadastrados",
     "1. Acessar listagem de produtos da loja id=1",
     "loja_id=1",
     "Exibidos: Iogurte Natural (R$ 1,99) e Queijo Fatiado (R$ 0,50)", "", "—", ""),

    # -----------------------------------------------------------------------
    # UC-04 — Gerir Vendas
    # -----------------------------------------------------------------------
    ("CT-023", "UC-04", "Listar vendas da loja", "Funcional",
     "Loja id=1 com pedido id=1 confirmado",
     "1. Acessar painel do lojista\n2. Acessar 'Minhas Vendas'",
     "loja_id=1",
     "Listagem exibe pedido id=1 com status 'Confirmado' e valor R$ 1,99", "", "—", ""),

    ("CT-024", "UC-04", "Filtrar vendas por data", "Funcional",
     "Loja id=1 com pedidos em datas distintas",
     "1. Acessar 'Minhas Vendas'\n2. Aplicar filtro de data: 2026-06-13",
     "loja_id=1 | data=2026-06-13",
     "Apenas pedidos da data filtrada são exibidos", "", "—", ""),

    ("CT-025", "UC-04", "Cancelar venda existente", "Funcional",
     "Pedido id=1 com status 'Confirmado'",
     "1. Selecionar pedido id=1\n2. Clicar em 'Cancelar'\n3. Confirmar",
     "pedido_id=1",
     "Status do pedido alterado para 'Cancelado'", "", "—", ""),

    ("CT-026", "UC-04", "Loja sem vendas exibe listagem vazia", "Funcional",
     "Loja recém-cadastrada sem pedidos",
     "1. Acessar 'Minhas Vendas' de loja sem pedidos",
     "loja_id=2",
     "Mensagem 'Nenhuma venda encontrada' exibida", "", "—", ""),

    # -----------------------------------------------------------------------
    # UC-05 — Manter Cadastro de Cliente
    # -----------------------------------------------------------------------
    ("CT-027", "UC-05", "Cadastrar cliente com dados válidos", "Funcional",
     "E-mail não cadastrado previamente",
     "1. Acessar 'Criar Cadastro'\n2. Preencher nome, e-mail, CPF e senha\n3. Confirmar",
     "nome=João Silva | email=joao@email.com | cpf=123.456.789-09 | senha=Senha@123",
     "Conta criada; usuário redirecionado para login ou home", "", "—", ""),

    ("CT-028", "UC-05", "Cadastrar cliente com e-mail já cadastrado", "Funcional",
     "E-mail 'teste@email.com' já existe na base",
     "1. Acessar 'Criar Cadastro'\n2. Inserir e-mail já existente\n3. Confirmar",
     "email=teste@email.com",
     "Sistema exibe erro 'E-mail já cadastrado'", "", "—", ""),

    ("CT-029", "UC-05", "Editar perfil do cliente", "Funcional",
     "Usuário id=1 autenticado",
     "1. Acessar 'Meu Perfil'\n2. Alterar nome\n3. Salvar",
     "id=1 | novo_nome=João Atualizado",
     "Dados atualizados e exibidos corretamente no perfil", "", "—", ""),

    ("CT-030", "UC-05", "Excluir conta de cliente", "Funcional",
     "Usuário id=1 autenticado; sem pedidos pendentes",
     "1. Acessar 'Meu Perfil'\n2. Clicar em 'Excluir Conta'\n3. Confirmar",
     "id=1",
     "Conta removida; sessão encerrada; redirecionado para tela de Login", "", "—", ""),

    ("CT-031", "UC-05", "Cadastrar cliente com CPF inválido", "Funcional",
     "Sistema disponível",
     "1. Acessar 'Criar Cadastro'\n2. Inserir CPF inválido\n3. Confirmar",
     "cpf=000.000.000-00",
     "Sistema exibe validação de CPF inválido", "", "—", ""),

    # -----------------------------------------------------------------------
    # UC-06 — Buscar por Lojas ou Alimentos
    # -----------------------------------------------------------------------
    ("CT-032", "UC-06", "Buscar produto por nome com resultado", "Funcional",
     "Produto 'Iogurte Natural' cadastrado",
     "1. Acessar campo de busca\n2. Digitar 'Iogurte'\n3. Confirmar busca",
     "termo=Iogurte",
     "Produto 'Iogurte Natural' exibido nos resultados", "", "—", ""),

    ("CT-033", "UC-06", "Buscar produto por nome sem resultado", "Funcional",
     "Nenhum produto com o nome buscado",
     "1. Acessar campo de busca\n2. Digitar 'Produto Inexistente'\n3. Confirmar busca",
     "termo=Produto Inexistente",
     "Mensagem 'Nenhum resultado encontrado' exibida", "", "—", ""),

    ("CT-034", "UC-06", "Buscar loja por nome", "Funcional",
     "Loja 'Mercado Teste' cadastrada e ativa",
     "1. Acessar campo de busca\n2. Digitar 'Mercado'\n3. Confirmar busca",
     "termo=Mercado",
     "Loja 'Mercado Teste' exibida nos resultados", "", "—", ""),

    ("CT-035", "UC-06", "Buscar com filtro de categoria", "Funcional",
     "Produtos de categorias distintas cadastrados",
     "1. Selecionar categoria 'Laticínios'\n2. Aplicar filtro",
     "categoria=Laticinios",
     "Apenas produtos da categoria 'Laticínios' exibidos", "", "—", ""),

    ("CT-036", "UC-06", "Busca retorna resultados em menos de 2 segundos", "Não Funcional",
     "Base com 100 produtos cadastrados",
     "1. Realizar busca\n2. Medir tempo de resposta",
     "termo=Iogurte",
     "Resultados exibidos em menos de 2000ms", "", "—", "Requisito de desempenho"),

    # -----------------------------------------------------------------------
    # UC-07 — Acessar Página de uma Loja
    # -----------------------------------------------------------------------
    ("CT-037", "UC-07", "Acessar página de loja ativa com produtos", "Funcional",
     "Loja id=1 ativa com 2 produtos",
     "1. Navegar para página da loja id=1",
     "loja_id=1",
     "Página exibe nome da loja, descrição e lista de produtos", "", "—", ""),

    ("CT-038", "UC-07", "Acessar página de loja sem produtos", "Funcional",
     "Loja id=1 ativa sem produtos cadastrados",
     "1. Navegar para página da loja id=1",
     "loja_id=1 (sem produtos)",
     "Página exibe mensagem 'Esta loja não possui produtos disponíveis'", "", "—", ""),

    ("CT-039", "UC-07", "Acessar página de loja inativa", "Funcional",
     "Loja id=1 com status Inativo",
     "1. Navegar para página da loja id=1",
     "loja_id=1 (inativa)",
     "Sistema exibe mensagem 'Loja indisponível' ou redireciona", "", "—", ""),

    ("CT-040", "UC-07", "Acessar página de loja inexistente", "Funcional",
     "Sistema disponível",
     "1. Navegar para página da loja id=999",
     "loja_id=999",
     "Sistema exibe erro 404 ou mensagem 'Loja não encontrada'", "", "—", ""),

    # -----------------------------------------------------------------------
    # UC-08 — Efetuar Compra
    # -----------------------------------------------------------------------
    ("CT-041", "UC-08", "Adicionar produto ao carrinho", "Funcional",
     "Usuário autenticado; produto id=1 disponível com estoque",
     "1. Acessar página da loja\n2. Clicar em 'Adicionar ao carrinho' no produto id=1",
     "produto_id=1 | quantidade=1",
     "Produto adicionado ao carrinho; ícone do carrinho atualizado", "", "—", ""),

    ("CT-042", "UC-08", "Finalizar compra com produto disponível", "Funcional",
     "Carrinho com produto id=1; usuário autenticado; método de pagamento configurado",
     "1. Acessar carrinho\n2. Revisar pedido\n3. Confirmar compra",
     "produto_id=1 | quantidade=1 | pagamento=cartao",
     "Pedido criado com status 'Confirmado'; e-mail de confirmação enviado", "", "—", ""),

    ("CT-043", "UC-08", "Tentar comprar produto sem estoque", "Funcional",
     "Produto id=1 com estoque=0",
     "1. Acessar produto sem estoque\n2. Tentar adicionar ao carrinho",
     "produto_id=1 | estoque=0",
     "Botão 'Adicionar ao carrinho' desativado ou sistema exibe 'Produto indisponível'", "", "—", ""),

    ("CT-044", "UC-08", "Remover produto do carrinho", "Funcional",
     "Carrinho com produto id=1",
     "1. Acessar carrinho\n2. Clicar em 'Remover' no produto id=1",
     "produto_id=1",
     "Produto removido; carrinho atualizado", "", "—", ""),

    ("CT-045", "UC-08", "Finalizar compra sem estar autenticado", "Funcional",
     "Usuário não autenticado",
     "1. Tentar finalizar compra sem login",
     "sessao=null",
     "Sistema redireciona para tela de Login", "", "—", ""),

    ("CT-046", "UC-08", "Processo de checkout completo < 5 segundos", "Não Funcional",
     "Usuário autenticado; carrinho com 1 produto",
     "1. Confirmar compra\n2. Medir tempo até confirmação",
     "produto_id=1",
     "Fluxo de checkout concluído em menos de 5000ms", "", "—", "Requisito de desempenho"),

    # -----------------------------------------------------------------------
    # UC-09 — Gerir Compras
    # -----------------------------------------------------------------------
    ("CT-047", "UC-09", "Visualizar histórico de compras", "Funcional",
     "Cliente id=1 com pedido id=1 confirmado",
     "1. Acessar 'Minhas Compras'",
     "cliente_id=1",
     "Pedido id=1 listado com status 'Confirmado' e valor R$ 1,99", "", "—", ""),

    ("CT-048", "UC-09", "Visualizar detalhes de um pedido", "Funcional",
     "Pedido id=1 existente para cliente id=1",
     "1. Acessar 'Minhas Compras'\n2. Clicar no pedido id=1",
     "pedido_id=1",
     "Detalhes exibidos: produto, quantidade, valor, data, status", "", "—", ""),

    ("CT-049", "UC-09", "Cancelar pedido com status 'Confirmado'", "Funcional",
     "Pedido id=1 com status 'Confirmado'",
     "1. Acessar detalhes do pedido id=1\n2. Clicar em 'Cancelar'\n3. Confirmar",
     "pedido_id=1 | status_atual=Confirmado",
     "Status alterado para 'Cancelado'; loja notificada", "", "—", ""),

    ("CT-050", "UC-09", "Tentar cancelar pedido já entregue", "Funcional",
     "Pedido id=1 com status 'Entregue'",
     "1. Acessar detalhes do pedido id=1\n2. Tentar cancelar",
     "pedido_id=1 | status_atual=Entregue",
     "Opção de cancelamento desativada; mensagem 'Pedido já entregue não pode ser cancelado'", "", "—", ""),

    ("CT-051", "UC-09", "Cliente sem histórico de compras", "Funcional",
     "Cliente id=2 sem pedidos realizados",
     "1. Acessar 'Minhas Compras' com conta sem pedidos",
     "cliente_id=2",
     "Mensagem 'Você ainda não realizou nenhuma compra' exibida", "", "—", ""),
]

# ---------------------------------------------------------------------------
# Dados: Massa de Testes
# ---------------------------------------------------------------------------
USUARIOS = [
    ("ID", "Nome", "E-mail", "Senha", "Perfil", "Status"),
    (1, "Teste", "teste@email.com", "123", "Lojista/Cliente", "Ativo"),
    (2, "Inativo", "inativo@email.com", "abc123", "Cliente", "Inativo"),
    (3, "João Silva", "joao@email.com", "Senha@123", "Cliente", "Ativo"),
]

LOJAS = [
    ("ID", "CNPJ", "Nome", "Descrição", "Dono (Usuario ID)", "Status"),
    (1, "12.345.678/0001-99", "Mercado Teste", "Loja de alimentos perto do prazo de validade", 1, "Ativo"),
    (2, "98.765.432/0001-11", "Padaria Silva", "Pães e frios com desconto", 3, "Inativo"),
]

PRODUTOS = [
    ("ID", "Nome", "Preço (R$)", "Data de Validade", "Estoque", "Loja ID", "Categoria", "Status"),
    (1, "Iogurte Natural", 1.99, "2026-06-16", 10, 1, "Laticínios", "Disponível"),
    (2, "Queijo Fatiado",  0.50, "2026-06-12", 5,  1, "Laticínios", "Vencido"),
    (3, "Pão Francês",     0.10, "2026-06-13", 50, 2, "Padaria",    "Disponível"),
]

PEDIDOS = [
    ("ID", "Cliente ID", "Loja ID", "Produto ID", "Quantidade", "Valor Total (R$)", "Data", "Status"),
    (1, 1, 1, 1, 1, 1.99, "2026-06-13", "Confirmado"),
    (2, 3, 1, 1, 2, 3.98, "2026-06-12", "Entregue"),
    (3, 1, 2, 3, 5, 0.50, "2026-06-11", "Cancelado"),
]

CENARIOS_NAO_FUNCIONAIS = [
    ("Requisito", "Métrica", "Valor Alvo", "Ferramenta Sugerida"),
    ("Tempo de resposta — Login", "Latência média", "< 2000ms", "JMeter / Postman"),
    ("Tempo de carregamento — Página de Loja", "Latência média", "< 3000ms", "Lighthouse / k6"),
    ("Tempo de busca", "Latência média", "< 2000ms", "k6 / Gatling"),
    ("Checkout completo", "Latência ponta a ponta", "< 5000ms", "Selenium + timestamp"),
    ("Concorrência", "Usuários simultâneos sem degradação", "≥ 50", "JMeter"),
    ("Disponibilidade", "Uptime mensal", "≥ 99%", "Monitoramento externo"),
]

# ---------------------------------------------------------------------------
# Construção do workbook
# ---------------------------------------------------------------------------
def criar_xlsx():
    wb = Workbook()

    # -----------------------------------------------------------------------
    # ABA 1 — Roteiro de Testes
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Roteiro de Testes"

    # Título principal
    ws1.merge_cells("A1:K1")
    titulo = ws1["A1"]
    titulo.value = "FS — Roteiro de Testes | Food Saver"
    titulo.fill = cabecalho_fill(COR_VERDE_ESCURO)
    titulo.font = Font(color="FFFFFF", bold=True, name="Calibri", size=13)
    titulo.alignment = alinhar("center")
    ws1.row_dimensions[1].height = 28

    # Subtítulo
    ws1.merge_cells("A2:K2")
    sub = ws1["A2"]
    sub.value = "Baseado nas Histórias de Usuário (UC-01 a UC-09) | Requisitos Funcionais e Não Funcionais"
    sub.fill = cabecalho_fill(COR_AZUL)
    sub.font = Font(color="FFFFFF", name="Calibri", size=10)
    sub.alignment = alinhar("center")
    ws1.row_dimensions[2].height = 18

    # Cabeçalho das colunas
    for col_idx, col_name in enumerate(COLUNAS_ROTEIRO, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=col_name)
        cell.fill = cabecalho_fill(COR_CINZA_CLARO)
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.alignment = alinhar("center", wrap=False)
        cell.border = borda
    ws1.row_dimensions[3].height = 18

    # Dados
    for row_idx, caso in enumerate(CASOS_TESTE, start=4):
        tipo = caso[3]
        fill_row = PatternFill("solid", fgColor=COR_VERDE_CLARO) if tipo == "Funcional" \
                   else PatternFill("solid", fgColor=COR_AMARELO)
        for col_idx, valor in enumerate(caso, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill_row
            cell.font = fonte_normal()
            cell.alignment = alinhar()
            cell.border = borda
        ws1.row_dimensions[row_idx].height = 60

    # Legenda
    row_leg = len(CASOS_TESTE) + 5
    ws1.merge_cells(f"A{row_leg}:B{row_leg}")
    ws1[f"A{row_leg}"].value = "Legenda de Tipo:"
    ws1[f"A{row_leg}"].font = Font(bold=True, name="Calibri", size=10)
    ws1[f"A{row_leg}"].alignment = alinhar()

    leg_f = ws1.cell(row=row_leg + 1, column=1, value="Funcional")
    leg_f.fill = PatternFill("solid", fgColor=COR_VERDE_CLARO)
    leg_f.font = fonte_normal()
    leg_f.border = borda
    leg_f.alignment = alinhar("center")

    leg_nf = ws1.cell(row=row_leg + 2, column=1, value="Não Funcional")
    leg_nf.fill = PatternFill("solid", fgColor=COR_AMARELO)
    leg_nf.font = fonte_normal()
    leg_nf.border = borda
    leg_nf.alignment = alinhar("center")

    # Larguras de coluna
    larguras = [10, 8, 38, 14, 35, 45, 38, 45, 25, 12, 20]
    for i, w in enumerate(larguras, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # -----------------------------------------------------------------------
    # ABA 2 — Massa de Testes
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Massa de Testes")

    row = 1

    def escrever_secao(titulo_secao, dados, cor_titulo=COR_VERDE_ESCURO):
        nonlocal row
        ncols = len(dados[0])
        ws2.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=ncols
        )
        cell_t = ws2.cell(row=row, column=1, value=titulo_secao)
        cell_t.fill = cabecalho_fill(cor_titulo)
        cell_t.font = fonte_branca_negrito()
        cell_t.alignment = alinhar("center", wrap=False)
        ws2.row_dimensions[row].height = 20
        row += 1

        for r_idx, linha in enumerate(dados):
            is_header = r_idx == 0
            for c_idx, valor in enumerate(linha, start=1):
                cell = ws2.cell(row=row, column=c_idx, value=valor)
                cell.font = Font(bold=is_header, name="Calibri", size=10)
                cell.alignment = alinhar()
                cell.border = borda
                if is_header:
                    cell.fill = cabecalho_fill(COR_CINZA_CLARO)
            ws2.row_dimensions[row].height = 18
            row += 1
        row += 1

    escrever_secao("USUÁRIOS — Massa de Testes", USUARIOS)
    escrever_secao("LOJAS — Massa de Testes", LOJAS, COR_AZUL)
    escrever_secao("PRODUTOS — Massa de Testes", PRODUTOS)
    escrever_secao("PEDIDOS — Massa de Testes", PEDIDOS, COR_AZUL)
    escrever_secao("REQUISITOS NÃO FUNCIONAIS — Referência", CENARIOS_NAO_FUNCIONAIS)

    for col_let in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws2.column_dimensions[col_let].width = 25

    # -----------------------------------------------------------------------
    # Salvar
    # -----------------------------------------------------------------------
    output_path = os.path.join(os.path.dirname(__file__), "FS - Roteiro de Testes.xlsx")
    wb.save(output_path)
    print(f"Arquivo gerado: {output_path}")


if __name__ == "__main__":
    criar_xlsx()
